#!/usr/bin/python
# -*- coding:utf-8 -*-
#
# filename: xlsx_parse.py
# Debug on Python 3.5.2
# 2017.04.17 by Oicebot 
#
# 运行方法，用 Python IDLE 打开，按 F5 运行
#
#

from openpyxl import Workbook, load_workbook
import parse_functions # parse_sonography, parse_pathology

table_title=['编号','姓名','年龄','住院号',
             '描述','位置','大小','形态','生长方向',
             '边缘','分布','内部回声','钙化','后方回声','CDFI'
             ]
             
table2_title = ['编号','姓名','年龄','住院号',
                '位置','定性','分级','单发/多灶',
                '伴随病变的情况','累及周围组织情况',
                '淋巴结转移情况'
             ]

def create_rows(indata_sheet, table_title, col_number=8, method='sonography', testrange=0):
    '''
    此函数用于读取具体工作表中所需的单元格数据，以 dict 对象的形式返回。
    
    indata_sheet ： 要处理的工作表对象
    col_number ： 要提取数据的列号
    method ： 提取方法： sonography, pathology ... etc
    testrange ： 提取多少层，默认 0 为不限

    '''
    
    row_objs = [] 
    
    index = 0
    
    #if testrange > 0:
    #    print("Parsing: {} Col: {} method: {} range: {} ".format(indata_sheet.title, col_number, method, testrange))
    
    for row in indata_sheet.rows:
        if testrange > 0:
            if index > testrange:
                break
    
        #函数化调用方法
        methodtocall = getattr(parse_functions,'parse_' + method)
        
        #if testrange > 0:
        #    print("Row {} : Col: {}, Value: {} ".format(index, col_number, row[col_number].value) )
         
        parsed_data = methodtocall(row[col_number].value)

        for data_item in parsed_data:

            info_data = dict.fromkeys(list(table_title),'')
            info_data['编号'] = row[0].value
            info_data['姓名'] = row[1].value
            info_data['年龄'] = row[3].value
            info_data['住院号'] = row[4].value
            
            for key,item in data_item.items():
                info_data[key] = item

            row_objs.append(info_data)
        index += 1

    return row_objs
    
    
def add_rows_to_table(display_info, table_title, indata_sheet, col=8, mtd='sonography', tstrange=0):
            
    outtable= [ table_title, ]

    row_objs = create_rows(indata_sheet, table_title, col_number=col, method=mtd, testrange=tstrange )
    
    index = 0
    print('添加{}：'.format(display_info))
    for i in row_objs:
        if index % 10 == 0:
            print(str(index),end='')
        else:
            print('.',end='')
        
        #print('----- {} -----'.format(index))
        index += 1
        next_row =[]
        for key in table_title:
            #print('{} : {}'.format(key,i[key]))
            next_row.append(i[key])

        outtable.append(next_row)
    print(' 共导出 {} 条数据。'.format(index))
    return outtable
        

if __name__ == '__main__':

    indata = load_workbook('testdata.xlsx')
    #indata = load_workbook('info.xlsx')
    outdata_filename = "outinfo.xlsx"

    print('该文件中含有如下工作表：')
    sheets = []
    for sheet in indata:
        sheets.append(sheet)

    print('===========')
    SheetID = 'A'
    indata_sheet = None
    while not SheetID.isdigit():
        
        i = 0
        for sheet in sheets:
            print ("{} : {}".format(i,sheet.title))
            i += 1
        print('请输入你想要解析的工作表的编号：',end='')
        SheetID = input()

        try:
            indata_sheet = sheets[int(SheetID)]
        except Exception as errinfo :
            print('打开失败……', errinfo)
            SheetID = 'A'
        else:
            print('已打开表：“{}”，解析中…'.format(indata_sheet.title))
            
    '''
    备注： indata_sheet的表头结构：
    0    1    2    3    4      5        6    7    8          9
    编号 姓名 性别 年龄 住院号 检查时间 病史 化疗 M-超声描述 M-超声诊断 ...
      14         15           21       22       23      24       25        26       27       28
    M-穿刺病理 M-穿刺免疫组化 钼靶描述 钼靶诊断 MRI描述 MRI诊断 手术名称 手术过程 术中冰冻 石蜡病理
    '''
            
    data1 = add_rows_to_table('描述数据', table_title, indata_sheet, col=8, mtd='sonography',)
    data2 = add_rows_to_table('病理诊断', table2_title, indata_sheet, col=28, mtd='pathology', tstrange=10, )

    wb = Workbook()
    ws = wb.active
    
    line_index = 0
    lastrow=[]
    current_side = -1   #0-left, 1-right 2-both
    current_ID = -1
    for row in data1:
        outrow = list(row)
        if line_index < len(data2):
            #第一行标题续后，之后只有编号相同的续上
            if line_index == 0 or data2[line_index][0] <= outrow[0]:

                while data2[line_index][0] < outrow[0]:             #如果切出来的东西比较多，就要塞空行进去
                    addrow = list(lastrow[:4])                      #前4列还是塞进去
                    addrow.extend(list(" " * 11))                   #塞入11列空格
                    addrow.extend(data2[line_index][4:])  #跳过 '编号','姓名','年龄','住院号',
                    ws.append(addrow)
                    line_index += 1

                if line_index > 0:                              #Exclude 1st/title row
                    if outrow[0] != current_ID:                 #Only change cache when meet new ID
                        current_ID = data2[line_index][0]       #cache current ID '编号'
                        #cache current side
                        if all(i in data2[line_index][4] for i in ['左','右']):
                            current_side = 2
                        elif '左' in data2[line_index][4]:
                            current_side = 0
                        elif '右' in data2[line_index][4]:
                            current_side = 1
                        else:
                            current_side = -1

                    compare_info = data2[line_index][4]
                    #if both sides exist, or left and right compared, then do nothing
                    if current_side == 2 or ('左' in outrow[4] and current_side == 0) or ('右' in outrow[4] and current_side == 1):
                        pass
                    else:
                        compare_info = compare_info + '【卧槽】'

                    outrow.append(compare_info)

                else:
                    outrow.append('位置')

                outrow.extend(data2[line_index][5:]) #跳过 '编号','姓名','年龄','住院号','位置'

                line_index += 1

            #//TODO: 石蜡病理的数量少， insert 【卧槽】 info
            else:
                compare_info = "" #data2[line_index][4]
                #if both sides exist, or left and right compared, then do nothing
                if current_side == 2 or ('左' in outrow[4] and current_side == 0) or ('右' in outrow[4] and current_side == 1):
                    pass
                else:
                    compare_info = compare_info + '【卧槽】'

                outrow.append(compare_info)
        
        ws.append(outrow)
        lastrow=list(outrow)  #保存之前处理过的行
        
    print('解析完毕！')

    wb.save(outdata_filename)

